import json
import hashlib
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from typing import Any

from app_paths import data_path, ensure_runtime_defaults
from app_logger import log_debug, log_warning

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from formula_parser import FormulaParser


ODS_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def _ods_attr(namespace: str, name: str) -> str:
    return f"{{{ODS_NS[namespace]}}}{name}"


class SimpleCell:
    def __init__(self, coordinate: str, value: Any):
        self.coordinate = coordinate
        self.value = value


class SimpleSheet:
    def __init__(self, title: str, rows: list[list[SimpleCell]]):
        self.title = title
        self._rows = rows

    def iter_rows(self):
        return iter(self._rows)


class SimpleWorkbook:
    def __init__(self, sheets: dict[str, SimpleSheet]):
        self._sheets = sheets
        self.sheetnames = list(sheets.keys())

    def __getitem__(self, sheet_name: str) -> SimpleSheet:
        return self._sheets[sheet_name]


class DataLoader:
    def __init__(self):
        ensure_runtime_defaults()
        self.workbook = None
        self.cell_cache: dict[str, dict[str, dict[str, Any]]] = {}
        self.app_meta: dict[str, Any] = {}
        self.active_cache_path = ""
        self.current_character_name = "unknown_character"
        self.source_file_path = ""
        self.is_dirty = False
        self.dirty_reason = ""
        self.last_loaded_snapshot_hash = ""
        self.last_saved_hash = ""
        self.load_cache_from_json()

    @property
    def active_character_path(self):
        return self.active_cache_path

    @active_character_path.setter
    def active_character_path(self, value):
        self.active_cache_path = value or ""

    def load_file(self, file_path):
        lower_path = str(file_path).lower()
        if lower_path.endswith(".ods"):
            self.workbook = self._load_ods_workbook(file_path)
        elif lower_path.endswith(".xlsx") or lower_path.endswith(".xlsm"):
            if load_workbook is None:
                raise ValueError("XLSX/XLSM benötigt das Python-Paket openpyxl.")
            self.workbook = load_workbook(file_path, data_only=False)
        else:
            raise ValueError("Dateityp nicht unterstützt. Bitte .xlsx, .xlsm oder .ods verwenden.")
        self.source_file_path = file_path
        self._build_cache()
        FormulaParser().recalculate_cache(self.cell_cache)
        self.save_cache_to_json()
        log_debug("cache", f"built: {len(self.cell_cache)}")

    def get_sheets(self):
        if self.cell_cache:
            return list(self.cell_cache.keys())
        if self.workbook is not None:
            return self.workbook.sheetnames
        return []

    def get_sheet_data(self, sheet_name):
        sheet_cache = self.cell_cache.get(sheet_name, {})
        if not sheet_cache:
            return []

        max_row = 0
        max_col = 0
        for cell_ref in sheet_cache.keys():
            row, col = self._cell_ref_to_row_col(cell_ref)
            if row > max_row:
                max_row = row
            if col > max_col:
                max_col = col

        data = []
        for row_index in range(1, max_row + 1):
            row_data = []
            for col_index in range(1, max_col + 1):
                cell_ref = self._col_to_letters(col_index) + str(row_index)
                cell_data = sheet_cache.get(cell_ref)
                row_data.append(cell_data.get("value") if cell_data else None)
            data.append(row_data)

        return data

    def get_sheet_object(self, sheet_name):
        if self.workbook is None:
            return None
        return self.workbook[sheet_name]

    def get_character_dir(self) -> str:
        path = data_path("characters")
        os.makedirs(str(path), exist_ok=True)
        return str(path)

    def get_legacy_cache_dir(self) -> str:
        path = data_path("cache")
        os.makedirs(str(path), exist_ok=True)
        return str(path)

    def save_cache_to_json(self, path=None):
        if path is None:
            character_name = self._get_character_name_from_cache()
            slug = self.make_character_slug(character_name)
            path = os.path.join(self.get_character_dir(), f"{slug}.json")
        return self.save_active_character_json(path)

    def save_active_character_json(self, path=None):
        if path is None:
            character_name = self._get_character_name_from_cache()
            slug = self.make_character_slug(character_name)
            path = self.active_cache_path or os.path.join(self.get_character_dir(), f"{slug}.json")
        else:
            character_name = self._get_character_name_from_cache()

        path = str(path)
        directory = os.path.dirname(path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        payload = self._to_serializable(
            {
                "cell_cache": self.cell_cache,
                "app_meta": self.app_meta,
            }
        )
        tmp_path = f"{path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
        self.active_cache_path = path
        self.current_character_name = character_name
        self._write_current_character_metadata(path, character_name, saved=True)
        self._remember_clean_snapshot()
        log_debug("cache", f"character: {character_name}")
        log_debug("cache", f"saved character cache: {path}")
        log_debug("cache", f"active: {path}")
        log_debug("save", f"active character saved: {path}")
        return True

    def load_cache_from_json(self, path=None):
        metadata_source_file = ""
        metadata_character_name = "unknown_character"
        if path is None:
            metadata_path = str(data_path("current_character.json"))
            if os.path.exists(metadata_path):
                try:
                    with open(metadata_path, "r", encoding="utf-8") as f:
                        metadata = json.load(f)
                    active_cache = metadata.get("active_cache")
                    if isinstance(active_cache, str) and active_cache:
                        path = active_cache
                        metadata_character_name = str(metadata.get("character_name", "unknown_character"))
                        metadata_source_file = str(metadata.get("source_file", ""))
                except Exception:
                    path = None

        if path is None:
            self.cell_cache = {}
            self.app_meta = {}
            self.active_cache_path = ""
            self.current_character_name = "unknown_character"
            self.source_file_path = ""
            self._remember_clean_snapshot()
            log_debug("cache", "no active character configured")
            return False
        if not os.path.exists(path):
            self.cell_cache = {}
            self.app_meta = {}
            self.active_cache_path = ""
            self.current_character_name = "unknown_character"
            self.source_file_path = ""
            self._remember_clean_snapshot()
            log_warning("cache", f"active character file missing: {path}")
            return False
        try:
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            self._load_cache_payload(loaded)
            self.active_cache_path = path
            detected_name = self._get_character_name_from_payload(loaded)
            self.current_character_name = (
                detected_name
                if detected_name != "unknown_character"
                else metadata_character_name
            )
            self.source_file_path = metadata_source_file or self.source_file_path
            self._remember_clean_snapshot()
            log_debug("cache", f"active: {path}")
            log_debug("cache", f"active character loaded: {path}")
            return True
        except Exception:
            self.cell_cache = {}
            self.app_meta = {}
            self.active_cache_path = ""
            self.current_character_name = "unknown_character"
            self._remember_clean_snapshot()
            log_warning("cache", f"active character file missing: {path}")
            return False

    def make_character_slug(self, name: str) -> str:
        if not isinstance(name, str):
            return "unknown_character"
        slug = name.strip().lower()
        slug = (
            slug.replace("ä", "ae")
            .replace("ö", "oe")
            .replace("ü", "ue")
            .replace("ß", "ss")
        )
        slug = re.sub(r"\s+", "_", slug)
        slug = re.sub(r"[^a-z0-9_]", "", slug)
        slug = re.sub(r"_+", "_", slug).strip("_")
        return slug if slug else "unknown_character"

    def list_character_caches(self) -> list[dict[str, str]]:
        results: list[dict[str, str]] = []
        seen_files = set()
        for storage, directory in (
            ("characters", self.get_character_dir()),
            ("legacy_cache", self.get_legacy_cache_dir()),
        ):
            if not os.path.isdir(directory):
                continue
            for file_name in sorted(os.listdir(directory)):
                if not file_name.lower().endswith(".json"):
                    continue
                lowered_file = file_name.lower()
                if lowered_file in seen_files:
                    continue
                seen_files.add(lowered_file)
                cache_path = os.path.join(directory, file_name)
                character_name = os.path.splitext(file_name)[0]
                character_format = "unknown"
                try:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        payload = json.load(f)
                    character_format = self.detect_character_json_format(payload)
                    detected_name = self._get_character_name_from_payload(payload)
                    if detected_name != "unknown_character":
                        character_name = detected_name
                except Exception:
                    pass
                results.append(
                    {
                        "name": character_name,
                        "path": cache_path,
                        "file": file_name,
                        "storage": storage,
                        "format": character_format,
                    }
                )
        return results

    def load_character_cache(self, cache_path: str) -> bool:
        if not cache_path or not os.path.exists(cache_path):
            return False
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if not isinstance(loaded, dict):
                return False
            character_format = self.detect_character_json_format(loaded)
            if character_format == "creator_state":
                self.cell_cache, self.app_meta = self.build_cache_from_creator_state(loaded)
                self.source_file_path = cache_path
                derived_path = self._next_creator_sheet_cache_path(cache_path)
                self.save_active_character_json(derived_path)
                log_debug("cache", f"creator character converted to sheet cache: {derived_path}")
                return True
            self._load_cache_payload(loaded)
            self.active_cache_path = cache_path
            self.current_character_name = self._get_character_name_from_cache()
            self._write_current_character_metadata(cache_path, self.current_character_name)
            self._remember_clean_snapshot()
            log_debug("cache", f"active: {cache_path}")
            log_debug("cache", f"active character loaded: {cache_path}")
            return True
        except Exception:
            return False

    def build_cache_from_creator_state(self, creator_state: dict) -> tuple[dict, dict]:
        creator_state = creator_state if isinstance(creator_state, dict) else {}
        species = creator_state.get("species", {}) if isinstance(creator_state.get("species"), dict) else {}
        concept = creator_state.get("concept", {}) if isinstance(creator_state.get("concept"), dict) else {}
        attributes = creator_state.get("attributes", {}) if isinstance(creator_state.get("attributes"), dict) else {}
        body = attributes.get("body", {}) if isinstance(attributes.get("body"), dict) else {}
        mind = attributes.get("mind", {}) if isinstance(attributes.get("mind"), dict) else {}
        skills = creator_state.get("skills", []) if isinstance(creator_state.get("skills"), list) else []
        perks = creator_state.get("perks", []) if isinstance(creator_state.get("perks"), list) else []
        equipment = creator_state.get("equipment", {}) if isinstance(creator_state.get("equipment"), dict) else {}
        money = equipment.get("money", {}) if isinstance(equipment.get("money"), dict) else {}
        items = equipment.get("items", []) if isinstance(equipment.get("items"), list) else []
        weapons = equipment.get("weapons", []) if isinstance(equipment.get("weapons"), list) else []
        armor = equipment.get("armor", []) if isinstance(equipment.get("armor"), list) else []

        cell_cache = {
            "Charakterbogen": {},
            "Fertigkeiten": {},
            "Inventar": {},
            "Ausrüstung": {},
            "Magie": {},
            "Notizen": {},
        }

        def set_cell(sheet_name, cell_ref, value):
            cell_cache.setdefault(sheet_name, {})[cell_ref] = self._creator_cell(value)

        character_name = self._clean_creator_text(concept.get("character_name"))
        set_cell("Charakterbogen", "G1", character_name)
        set_cell("Charakterbogen", "C1", character_name)
        set_cell("Charakterbogen", "G3", self._clean_creator_text(species.get("name")))
        set_cell("Charakterbogen", "G5", self._clean_creator_text(concept.get("short_concept")))
        set_cell("Charakterbogen", "G7", self._clean_creator_text(concept.get("origin")))
        set_cell("Charakterbogen", "G9", self._clean_creator_text(concept.get("role")))
        set_cell("Charakterbogen", "AG7", self._creator_int(body.get("kraft")))
        set_cell("Charakterbogen", "AG9", self._creator_int(body.get("geschick")))
        set_cell("Charakterbogen", "AG11", self._creator_int(body.get("zaehigkeit")))
        set_cell("Charakterbogen", "AG13", self._creator_int(body.get("reflex")))
        set_cell("Charakterbogen", "AR7", self._creator_int(mind.get("intelligenz")))
        set_cell("Charakterbogen", "AR9", self._creator_int(mind.get("willenskraft")))
        set_cell("Charakterbogen", "AR11", self._creator_int(mind.get("charisma")))
        set_cell("Charakterbogen", "AR13", self._creator_int(mind.get("sinne")))

        set_cell("Fertigkeiten", "A1", "Fertigkeiten")
        set_cell("Fertigkeiten", "A2", "Name")
        set_cell("Fertigkeiten", "B2", "Attribut")
        set_cell("Fertigkeiten", "C2", "Spezialisierung")
        set_cell("Fertigkeiten", "D2", "BP")
        for index, skill in enumerate(skills, start=3):
            if not isinstance(skill, dict):
                continue
            set_cell("Fertigkeiten", f"A{index}", self._clean_creator_text(skill.get("name") or skill.get("id")))
            set_cell("Fertigkeiten", f"B{index}", self._clean_creator_text(skill.get("attribute")))
            set_cell("Fertigkeiten", f"C{index}", self._clean_creator_text(skill.get("specialization")))
            set_cell("Fertigkeiten", f"D{index}", self._creator_int(skill.get("bp")))

        set_cell("Charakterbogen", "A30", "Perks")
        for index, perk in enumerate(perks, start=31):
            if not isinstance(perk, dict):
                continue
            set_cell("Charakterbogen", f"A{index}", self._clean_creator_text(perk.get("name")))
            set_cell("Charakterbogen", f"B{index}", self._creator_int(perk.get("bp")))
            set_cell("Charakterbogen", f"C{index}", self._clean_creator_text(perk.get("effect")))

        set_cell("Inventar", "B9", self._creator_int(money.get("gulden")))
        set_cell("Inventar", "E9", self._creator_int(money.get("schilling")))
        set_cell("Inventar", "H9", self._creator_int(money.get("heller")))
        set_cell("Inventar", "A12", "Inventar")
        set_cell("Inventar", "B12", "PL")
        set_cell("Inventar", "C12", "Anzahl")
        for index, item in enumerate(items, start=13):
            if not isinstance(item, dict):
                continue
            set_cell("Inventar", f"A{index}", self._clean_creator_text(item.get("name")))
            set_cell("Inventar", f"B{index}", self._clean_creator_text(item.get("pl")))
            set_cell("Inventar", f"C{index}", self._clean_creator_text(item.get("count")))

        set_cell("Ausrüstung", "A1", "Waffen")
        set_cell("Ausrüstung", "A2", "Name")
        set_cell("Ausrüstung", "B2", "Schaden / Effekt")
        set_cell("Ausrüstung", "C2", "Attribut / Fertigkeit")
        set_cell("Ausrüstung", "D2", "Notiz")
        for index, weapon in enumerate(weapons, start=3):
            if not isinstance(weapon, dict):
                continue
            set_cell("Ausrüstung", f"A{index}", self._clean_creator_text(weapon.get("name")))
            set_cell("Ausrüstung", f"B{index}", self._clean_creator_text(weapon.get("damage")))
            set_cell("Ausrüstung", f"C{index}", self._clean_creator_text(weapon.get("attribute")))
            set_cell("Ausrüstung", f"D{index}", self._clean_creator_text(weapon.get("note")))
        set_cell("Ausrüstung", "A20", "Rüstung")
        set_cell("Ausrüstung", "A21", "Name")
        set_cell("Ausrüstung", "B21", "Schutz / Werte")
        set_cell("Ausrüstung", "C21", "Notiz")
        for index, armor_row in enumerate(armor, start=22):
            if not isinstance(armor_row, dict):
                continue
            set_cell("Ausrüstung", f"A{index}", self._clean_creator_text(armor_row.get("name")))
            set_cell("Ausrüstung", f"B{index}", self._clean_creator_text(armor_row.get("protection")))
            set_cell("Ausrüstung", f"C{index}", self._clean_creator_text(armor_row.get("note")))

        set_cell("Notizen", "A1", "Notizen")
        set_cell("Notizen", "A2", self._clean_creator_text(concept.get("description")))
        set_cell("Notizen", "A4", self._clean_creator_text(concept.get("motivation")))
        set_cell("Notizen", "A6", self._clean_creator_text(equipment.get("notes")))

        app_meta = {
            "creator_state": creator_state,
            "source_format": "creator_state",
            "created_by": "character_creator",
            "concept": concept,
            "species": species,
            "bp": creator_state.get("bp", {}) if isinstance(creator_state.get("bp"), dict) else {},
            "creator_skills": skills,
            "creator_perks": perks,
            "creator_equipment": equipment,
            "creator_notes": {
                "description": self._clean_creator_text(concept.get("description")),
                "motivation": self._clean_creator_text(concept.get("motivation")),
                "equipment": self._clean_creator_text(equipment.get("notes")),
            },
        }
        return cell_cache, app_meta

    def mark_dirty(self, reason=""):
        self.is_dirty = True
        self.dirty_reason = str(reason or "")
        log_debug("save", f"character dirty: {self.dirty_reason}")

    def has_unsaved_changes(self):
        if self.is_dirty:
            return True
        current_hash = self._snapshot_hash()
        saved_hash = self.last_saved_hash or self.last_loaded_snapshot_hash
        return bool(saved_hash and current_hash != saved_hash)

    def set_cell_value(self, sheet_name, cell_ref, value, mark_dirty=True):
        if not sheet_name or not cell_ref:
            return False
        sheet_cache = self.cell_cache.setdefault(sheet_name, {})
        cell_data = sheet_cache.get(cell_ref)
        if not isinstance(cell_data, dict):
            cell_data = {
                "value": None,
                "formula": None,
                "references": [],
                "error": None,
            }
            sheet_cache[cell_ref] = cell_data

        cell_data["value"] = value
        cell_data.setdefault("formula", None)
        cell_data.setdefault("references", [])
        cell_data.setdefault("error", None)
        log_debug("save", f"{sheet_name}!{cell_ref} = {value}")
        if mark_dirty:
            self.mark_dirty(f"{sheet_name}!{cell_ref}")
        return True

    def _build_cache(self):
        if self.workbook is None:
            self.cell_cache = {}
            return
        self.cell_cache = {}
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.cell_cache[sheet_name] = {}
            for row in sheet.iter_rows():
                for cell in row:
                    raw_value = cell.value
                    formula = raw_value if isinstance(raw_value, str) and raw_value.startswith("=") else None
                    references = self._extract_references(formula) if formula else []
                    if raw_value is None and formula is None and not references:
                        continue
                    self.cell_cache[sheet_name][cell.coordinate] = {
                        "value": raw_value,
                        "formula": formula,
                        "references": references,
                        "error": None,
                    }

    def _load_ods_workbook(self, file_path) -> SimpleWorkbook:
        try:
            with zipfile.ZipFile(file_path) as ods:
                content = ods.read("content.xml")
        except (KeyError, zipfile.BadZipFile) as exc:
            raise ValueError("ODS-Datei konnte nicht gelesen werden.") from exc

        root = ET.fromstring(content)
        spreadsheet = root.find(".//office:spreadsheet", ODS_NS)
        if spreadsheet is None:
            raise ValueError("ODS-Datei enthält keine Tabellen.")

        sheets: dict[str, SimpleSheet] = {}
        for table_index, table in enumerate(spreadsheet.findall("table:table", ODS_NS), start=1):
            sheet_name = table.get(_ods_attr("table", "name")) or f"Sheet{table_index}"
            rows: list[list[SimpleCell]] = []
            row_index = 1

            for row_element in table.iter(_ods_attr("table", "table-row")):
                row_repeat = self._ods_repeat(row_element, "number-rows-repeated")
                row_cells, has_content = self._parse_ods_row(row_element, row_index)
                if has_content:
                    rows.append(row_cells)
                    for offset in range(1, min(row_repeat, 1024)):
                        rows.append(
                            [
                                SimpleCell(
                                    f"{self._col_to_letters(col_index)}{row_index + offset}",
                                    cell.value,
                                )
                                for col_index, cell in enumerate(row_cells, start=1)
                            ]
                        )
                row_index += row_repeat

            sheets[sheet_name] = SimpleSheet(sheet_name, rows)

        if not sheets:
            raise ValueError("ODS-Datei enthält keine Tabellen.")
        log_debug("cache", f"ODS loaded: {file_path}")
        return SimpleWorkbook(sheets)

    def _parse_ods_row(self, row_element, row_index: int) -> tuple[list[SimpleCell], bool]:
        segments: list[tuple[int, Any]] = []
        col_index = 1
        max_content_col = 0

        cell_tags = {
            _ods_attr("table", "table-cell"),
            _ods_attr("table", "covered-table-cell"),
        }
        for cell_element in row_element:
            if cell_element.tag not in cell_tags:
                continue
            repeat = self._ods_repeat(cell_element, "number-columns-repeated")
            cell_value = (
                None
                if cell_element.tag == _ods_attr("table", "covered-table-cell")
                else self._ods_cell_value(cell_element)
            )
            segments.append((repeat, cell_value))
            if cell_value is not None:
                max_content_col = col_index + repeat - 1
            col_index += repeat

        if max_content_col == 0:
            return [], False

        row_cells: list[SimpleCell] = []
        col_index = 1
        for repeat, cell_value in segments:
            for _ in range(repeat):
                if col_index > max_content_col:
                    break
                row_cells.append(
                    SimpleCell(f"{self._col_to_letters(col_index)}{row_index}", cell_value)
                )
                col_index += 1
            if col_index > max_content_col:
                break

        return row_cells, True

    def _ods_cell_value(self, cell_element):
        formula = cell_element.get(_ods_attr("table", "formula"))
        if formula:
            return self._normalize_ods_formula(formula)

        value_type = cell_element.get(_ods_attr("office", "value-type"))
        if value_type in {"float", "percentage", "currency"}:
            value = cell_element.get(_ods_attr("office", "value"))
            if value is None:
                return None
            try:
                number = float(value)
                return int(number) if number.is_integer() else number
            except ValueError:
                return value

        if value_type == "boolean":
            return cell_element.get(_ods_attr("office", "boolean-value")) == "true"

        if value_type == "date":
            return cell_element.get(_ods_attr("office", "date-value"))

        if value_type == "time":
            return cell_element.get(_ods_attr("office", "time-value"))

        text_parts = [
            "".join(paragraph.itertext())
            for paragraph in cell_element.findall("text:p", ODS_NS)
        ]
        text_value = "\n".join(part for part in text_parts if part != "")
        return text_value if text_value != "" else None

    def _normalize_ods_formula(self, formula: str) -> str:
        expression = formula.strip()
        if ":=" in expression:
            expression = expression.split(":=", 1)[1]
        elif expression.startswith("="):
            expression = expression[1:]

        def replace_bracket_ref(match):
            ref = match.group(1).replace("$", "")
            if ref.startswith("."):
                return ref[1:].replace(":.", ":")
            if "." in ref:
                sheet_name, cell_ref = ref.rsplit(".", 1)
                return f"{sheet_name}!{cell_ref}"
            return ref

        expression = re.sub(r"\[([^\]]+)\]", replace_bracket_ref, expression)
        return "=" + expression

    def _ods_repeat(self, element, attr_name: str) -> int:
        raw_repeat = element.get(_ods_attr("table", attr_name), "1")
        try:
            repeat = int(raw_repeat)
        except (TypeError, ValueError):
            repeat = 1
        return max(repeat, 1)

    def _extract_references(self, formula):
        refs = re.findall(r"[A-Za-z]+[0-9]+", formula or "")
        unique = []
        seen = set()
        for ref in refs:
            ref_up = ref.upper()
            if ref_up not in seen:
                seen.add(ref_up)
                unique.append(ref_up)
        return unique

    def _cell_ref_to_row_col(self, cell_ref):
        match = re.fullmatch(r"([A-Za-z]+)([0-9]+)", cell_ref)
        if not match:
            return (0, 0)
        letters = match.group(1).upper()
        row = int(match.group(2))
        col = 0
        for letter in letters:
            col = col * 26 + (ord(letter) - ord("A") + 1)
        return (row, col)

    def _col_to_letters(self, col):
        letters = ""
        number = col
        while number > 0:
            number, remainder = divmod(number - 1, 26)
            letters = chr(ord("A") + remainder) + letters
        return letters

    def _to_serializable(self, value):
        if isinstance(value, dict):
            return {k: self._to_serializable(v) for k, v in value.items()}
        if isinstance(value, list):
            return [self._to_serializable(v) for v in value]
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    def _creator_cell(self, value):
        return {
            "value": value,
            "formula": None,
            "references": [],
            "error": None,
        }

    def _clean_creator_text(self, value):
        return str(value or "").strip()

    def _creator_int(self, value):
        try:
            return int(value or 0)
        except Exception:
            return 0

    def _next_creator_sheet_cache_path(self, creator_path):
        directory = os.path.dirname(str(creator_path)) or self.get_character_dir()
        base_name = os.path.splitext(os.path.basename(str(creator_path)))[0] or "unnamed_character"
        sheet_base = f"{base_name}_sheet"
        candidate = os.path.join(directory, f"{sheet_base}.json")
        if not os.path.exists(candidate):
            return candidate
        index = 2
        while True:
            candidate = os.path.join(directory, f"{sheet_base}_{index}.json")
            if not os.path.exists(candidate):
                return candidate
            index += 1

    def _snapshot_hash(self, payload=None):
        if payload is None:
            payload = {"cell_cache": self.cell_cache, "app_meta": self.app_meta}
        serializable = self._to_serializable(payload)
        encoded = json.dumps(
            serializable,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    def _remember_clean_snapshot(self):
        snapshot_hash = self._snapshot_hash()
        self.last_loaded_snapshot_hash = snapshot_hash
        self.last_saved_hash = snapshot_hash
        self.is_dirty = False
        self.dirty_reason = ""

    def _get_character_name_from_cache(self) -> str:
        return self._get_character_name_from_payload(self.cell_cache)

    def _get_character_name_from_payload(self, payload) -> str:
        if not isinstance(payload, dict):
            return "unknown_character"
        concept = payload.get("concept")
        if isinstance(concept, dict):
            character_name = concept.get("character_name")
            if self._looks_like_character_name(character_name):
                return character_name.strip()
        if isinstance(payload.get("cell_cache"), dict):
            payload = payload.get("cell_cache", {})

        sheet_candidates = ["Charakterbogen", "CharacterSheet", "Sheet1"]
        for sheet_name in sheet_candidates:
            sheet_cache = payload.get(sheet_name)
            name = self._get_character_name_from_sheet_cache(sheet_cache)
            if name != "unknown_character":
                return name

        for sheet_cache in payload.values():
            name = self._get_character_name_from_sheet_cache(sheet_cache)
            if name != "unknown_character":
                return name
        return "unknown_character"

    def _get_character_name_from_sheet_cache(self, sheet_cache) -> str:
        if not isinstance(sheet_cache, dict):
            return "unknown_character"

        for cell_ref in ("G1", "C1"):
            value = self._get_cached_cell_value(sheet_cache, cell_ref)
            if self._looks_like_character_name(value):
                return value.strip()

        for cell_ref, cell_data in sheet_cache.items():
            value = cell_data.get("value") if isinstance(cell_data, dict) else None
            if not isinstance(value, str) or value.strip().lower() != "name":
                continue
            row, col = self._cell_ref_to_row_col(cell_ref)
            for next_col in range(col + 1, col + 8):
                next_ref = self._col_to_letters(next_col) + str(row)
                next_value = self._get_cached_cell_value(sheet_cache, next_ref)
                if self._looks_like_character_name(next_value):
                    return next_value.strip()

        return "unknown_character"

    def _get_cached_cell_value(self, sheet_cache, cell_ref: str):
        cell_data = sheet_cache.get(cell_ref)
        if isinstance(cell_data, dict):
            return cell_data.get("value")
        return None

    def _looks_like_character_name(self, value) -> bool:
        if not isinstance(value, str):
            return False
        text = value.strip()
        if not text:
            return False
        return text.lower() not in {"name", "unknown_character", "attribute"}

    def detect_character_json_format(self, payload) -> str:
        if not isinstance(payload, dict):
            return "unknown"
        if isinstance(payload.get("cell_cache"), dict):
            return "sheet_cache"
        if (
            "version" in payload
            and isinstance(payload.get("species"), dict)
            and isinstance(payload.get("concept"), dict)
            and isinstance(payload.get("attributes"), dict)
            and isinstance(payload.get("equipment"), dict)
        ):
            return "creator_state"
        if payload and all(isinstance(value, dict) for value in payload.values()):
            for sheet_cache in payload.values():
                if any(isinstance(cell_data, dict) and "value" in cell_data for cell_data in sheet_cache.values()):
                    return "legacy_cell_cache"
        return "unknown"

    def _load_cache_payload(self, payload):
        if not isinstance(payload, dict):
            self.cell_cache = {}
            self.app_meta = {}
            return
        if self.detect_character_json_format(payload) == "creator_state":
            self.cell_cache = {}
            self.app_meta = {
                "creator_state": payload,
                "source_format": "creator_state",
            }
            return
        if isinstance(payload.get("cell_cache"), dict):
            self.cell_cache = payload.get("cell_cache", {})
            app_meta = payload.get("app_meta", {})
            self.app_meta = app_meta if isinstance(app_meta, dict) else {}
            return
        self.cell_cache = payload
        self.app_meta = {}

    def get_inventory_tab_labels(self) -> dict[str, str]:
        inventory_meta = self.app_meta.get("inventory", {})
        if not isinstance(inventory_meta, dict):
            return {}
        labels = inventory_meta.get("tab_labels", {})
        if not isinstance(labels, dict):
            return {}
        result = {}
        for key, value in labels.items():
            if not isinstance(key, str):
                continue
            result[key] = str(value)
        return result

    def set_inventory_tab_label(self, slot_id: str, label: str):
        slot_key = str(slot_id or "").strip()
        if not slot_key:
            return False
        inventory_meta = self.app_meta.setdefault("inventory", {})
        if not isinstance(inventory_meta, dict):
            inventory_meta = {}
            self.app_meta["inventory"] = inventory_meta
        tab_labels = inventory_meta.setdefault("tab_labels", {})
        if not isinstance(tab_labels, dict):
            tab_labels = {}
            inventory_meta["tab_labels"] = tab_labels
        tab_labels[slot_key] = str(label)
        return True

    def get_inventory_custom_rows(self, slot_id: str) -> list[dict[str, str]]:
        inventory_meta = self.app_meta.get("inventory", {})
        if not isinstance(inventory_meta, dict):
            return []
        custom_rows = inventory_meta.get("custom_rows", {})
        if not isinstance(custom_rows, dict):
            return []
        rows = custom_rows.get(str(slot_id or "").strip(), [])
        if not isinstance(rows, list):
            return []
        result = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            result.append(
                {
                    "name": str(row.get("name", "") or ""),
                    "pl": str(row.get("pl", "") or ""),
                    "count": str(row.get("count", "") or ""),
                }
            )
        return result

    def set_inventory_custom_row_value(self, slot_id: str, row_index: int, field: str, value: str):
        slot_key = str(slot_id or "").strip()
        if not slot_key:
            return False
        field_key = str(field or "").strip()
        if field_key not in {"name", "pl", "count"}:
            return False
        try:
            index = int(row_index)
        except Exception:
            return False
        if index < 0:
            return False
        inventory_meta = self.app_meta.setdefault("inventory", {})
        if not isinstance(inventory_meta, dict):
            inventory_meta = {}
            self.app_meta["inventory"] = inventory_meta
        custom_rows = inventory_meta.setdefault("custom_rows", {})
        if not isinstance(custom_rows, dict):
            custom_rows = {}
            inventory_meta["custom_rows"] = custom_rows
        rows = custom_rows.setdefault(slot_key, [])
        if not isinstance(rows, list):
            rows = []
            custom_rows[slot_key] = rows
        while len(rows) <= index:
            rows.append({"name": "", "pl": "", "count": ""})
        if not isinstance(rows[index], dict):
            rows[index] = {"name": "", "pl": "", "count": ""}
        for key in ("name", "pl", "count"):
            rows[index].setdefault(key, "")
        rows[index][field_key] = str(value)
        return True

    def _write_current_character_metadata(self, active_cache_path: str, character_name: str, saved: bool = False) -> None:
        metadata_path = str(data_path("current_character.json"))
        os.makedirs(os.path.dirname(metadata_path), exist_ok=True)
        now = datetime.now().isoformat()
        payload = {
            "active_cache": active_cache_path,
            "active_character": active_cache_path,
            "character_name": character_name,
            "source_file": self.source_file_path,
            "last_loaded": now,
        }
        if saved:
            payload["last_saved"] = now
        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
